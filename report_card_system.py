#!/usr/bin/env python3
"""
Uganda Secondary School Automatic Report Card System
Aligned with NCDC / UNEB Competency-Based Curriculum (CBC)
Lower Secondary (S1–S4 / UCE)

Features:
- 20% Formative (Continuous Assessment / AOIs) + 80% Summative
- Official A–E grading with descriptors
- Batch generation of professional PDF report cards from Excel
- Class summary sheets
- Configurable school details
"""

import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm, cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    Image, KeepTogether, HRFlowable, PageBreak
)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# ============================================================
# CONFIGURATION – Edit these for your school
# ============================================================

SCHOOL_CONFIG = {
    "name": "ST. MARY'S SECONDARY SCHOOL",
    "motto": "Knowledge is Power",
    "address": "P.O. Box 1234, Kampala, Uganda",
    "phone": "+256 700 000000",
    "email": "info@stmaryss.ac.ug",
    "district": "Kampala",
    "logo_path": None,  # Put path to school logo PNG/JPG if available
}

# Official CBC Grading Scale (UNEB / NCDC)
GRADE_SCALE = [
    (80, 100, "A", "Exceptional",
     "Demonstrates extraordinary competency by applying knowledge and skills innovatively and creatively in real-life situations."),
    (70, 79, "B", "Outstanding",
     "Demonstrates a high level of competency by effectively applying acquired knowledge and skills in real-life situations."),
    (60, 69, "C", "Satisfactory",
     "Demonstrates an adequate level of competency in applying knowledge and skills in real-life situations."),
    (50, 59, "D", "Basic",
     "Demonstrates a minimum level of competency in applying knowledge and skills in real-life situations."),
    (0, 49, "E", "Elementary",
     "Demonstrates below the basic level of competency in applying knowledge and skills in real-life situations."),
]

# Common subjects for Lower Secondary (can be extended)
DEFAULT_SUBJECTS = [
    "English", "Mathematics", "Biology", "Chemistry", "Physics",
    "Geography", "History & Political Education", "Kiswahili",
    "Religious Education", "Entrepreneurship", "Physical Education",
    "ICT", "Agriculture", "Art and Design"
]


def get_grade(score: float) -> Tuple[str, str, str]:
    """Return (letter, descriptor, full_description) for a given score."""
    if score is None or pd.isna(score):
        return ("–", "Not Assessed", "")
    score = float(score)
    for low, high, letter, desc, full in GRADE_SCALE:
        if low <= score <= high:
            return (letter, desc, full)
    return ("E", "Elementary", GRADE_SCALE[-1][4])


def calculate_final_score(formative: float, summative: float) -> Optional[float]:
    """20% formative + 80% summative."""
    if pd.isna(formative) and pd.isna(summative):
        return None
    f = float(formative) if not pd.isna(formative) else 0.0
    s = float(summative) if not pd.isna(summative) else 0.0
    return round((f * 0.20) + (s * 0.80), 1)


# ============================================================
# EXCEL TEMPLATE GENERATOR
# ============================================================

def create_excel_template(output_path: str = "templates/marks_entry_template.xlsx"):
    """Create a ready-to-use Excel template for teachers."""
    wb = Workbook()

    # ----- Sheet 1: School Settings -----
    ws_settings = wb.active
    ws_settings.title = "School Settings"

    settings_data = [
        ["SCHOOL REPORT CARD SETTINGS", ""],
        ["", ""],
        ["School Name", SCHOOL_CONFIG["name"]],
        ["Motto", SCHOOL_CONFIG["motto"]],
        ["Address", SCHOOL_CONFIG["address"]],
        ["Phone", SCHOOL_CONFIG["phone"]],
        ["Email", SCHOOL_CONFIG["email"]],
        ["District", SCHOOL_CONFIG["district"]],
        ["", ""],
        ["Term", "Term 3"],
        ["Academic Year", "2025/2026"],
        ["Report Date", datetime.now().strftime("%d/%m/%Y")],
        ["", ""],
        ["INSTRUCTIONS", ""],
        ["1. Fill the 'Students' sheet with learner details.", ""],
        ["2. Fill the 'Marks' sheet with Formative (out of 100) and Summative (out of 100) scores.", ""],
        ["3. Leave blank if a subject was not taken by the learner.", ""],
        ["4. Run the system: python report_card_system.py --excel templates/marks_entry_template.xlsx", ""],
        ["5. Generated PDFs will appear in the 'generated_reports' folder.", ""],
    ]
    for row in settings_data:
        ws_settings.append(row)

    ws_settings["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws_settings["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws_settings.merge_cells("A1:B1")
    ws_settings.column_dimensions["A"].width = 25
    ws_settings.column_dimensions["B"].width = 55

    # ----- Sheet 2: Students -----
    ws_students = wb.create_sheet("Students")
    headers = [
        "Admission No", "Surname", "First Name", "Other Names", "Sex",
        "Class", "Stream", "Date of Birth", "Parent/Guardian", "Contact",
        "Attendance (%)", "Days Present", "Days Absent", "Class Teacher Comment",
        "Head Teacher Comment"
    ]
    ws_students.append(headers)

    # Sample students
    samples = [
        ["S2024/001", "Nakato", "Amina", "", "F", "S3", "East", "15/03/2009",
         "Mrs. Nakato Sarah", "0700123456", 92, 55, 5,
         "Amina is a hardworking and respectful learner. She participates actively in class discussions and group projects.",
         "Promoted to the next class. Keep up the excellent work."],
        ["S2024/015", "Okello", "David", "James", "M", "S3", "East", "22/07/2008",
         "Mr. Okello Peter", "0777987654", 85, 51, 9,
         "David shows good understanding in sciences but needs to improve consistency in Mathematics and English.",
         "Promoted. More effort required in core subjects."],
        ["S2024/028", "Namukasa", "Grace", "", "F", "S2", "West", "10/11/2010",
         "Mrs. Namukasa Rose", "0755112233", 96, 58, 2,
         "Grace is an outstanding all-rounder. Excellent in both academics and co-curricular activities.",
         "Excellent performance. Continue to be a role model."],
    ]
    for s in samples:
        ws_students.append(s)

    for col in range(1, len(headers) + 1):
        cell = ws_students.cell(1, col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(wrap_text=True, horizontal="center")
    ws_students.column_dimensions["A"].width = 14
    ws_students.column_dimensions["B"].width = 14
    ws_students.column_dimensions["C"].width = 12
    ws_students.column_dimensions["N"].width = 50
    ws_students.column_dimensions["O"].width = 40

    # ----- Sheet 3: Marks -----
    ws_marks = wb.create_sheet("Marks")
    mark_headers = ["Admission No", "Subject", "Formative (out of 100)", "Summative (out of 100)", "Teacher Comment (optional)"]
    ws_marks.append(mark_headers)

    # Sample marks for the three students
    sample_marks = [
        # Amina Nakato S3
        ["S2024/001", "English", 78, 82, "Good command of language"],
        ["S2024/001", "Mathematics", 65, 71, "Needs more practice on algebra"],
        ["S2024/001", "Biology", 85, 88, "Excellent practical skills"],
        ["S2024/001", "Chemistry", 72, 76, "Good understanding of concepts"],
        ["S2024/001", "Physics", 68, 74, "Improving steadily"],
        ["S2024/001", "Geography", 80, 84, "Very good map work"],
        ["S2024/001", "History & Political Education", 75, 79, "Good analysis"],
        ["S2024/001", "Religious Education", 88, 90, "Outstanding"],
        ["S2024/001", "Entrepreneurship", 82, 85, "Creative ideas"],
        ["S2024/001", "ICT", 90, 92, "Excellent digital skills"],
        # David Okello
        ["S2024/015", "English", 55, 62, "Needs improvement in composition"],
        ["S2024/015", "Mathematics", 48, 55, "Struggles with problem solving"],
        ["S2024/015", "Biology", 70, 75, "Good effort"],
        ["S2024/015", "Chemistry", 68, 72, "Satisfactory"],
        ["S2024/015", "Physics", 62, 68, "Can do better"],
        ["S2024/015", "Geography", 58, 65, "Average performance"],
        ["S2024/015", "History & Political Education", 60, 66, "Fair"],
        ["S2024/015", "Religious Education", 72, 78, "Good"],
        ["S2024/015", "Entrepreneurship", 65, 70, "Shows interest"],
        ["S2024/015", "ICT", 75, 80, "Competent"],
        # Grace Namukasa S2
        ["S2024/028", "English", 88, 91, "Excellent"],
        ["S2024/028", "Mathematics", 92, 95, "Top performer"],
        ["S2024/028", "Biology", 85, 89, "Very strong"],
        ["S2024/028", "Chemistry", 80, 86, "Good practicals"],
        ["S2024/028", "Physics", 78, 84, "Solid understanding"],
        ["S2024/028", "Geography", 90, 93, "Outstanding"],
        ["S2024/028", "History & Political Education", 82, 88, "Very good"],
        ["S2024/028", "Kiswahili", 75, 80, "Good progress"],
        ["S2024/028", "Religious Education", 95, 97, "Exceptional"],
        ["S2024/028", "Entrepreneurship", 88, 90, "Innovative"],
        ["S2024/028", "Physical Education", 90, 92, "Active and disciplined"],
        ["S2024/028", "ICT", 94, 96, "Excellent"],
    ]
    for m in sample_marks:
        ws_marks.append(m)

    for col in range(1, 6):
        cell = ws_marks.cell(1, col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
    ws_marks.column_dimensions["A"].width = 14
    ws_marks.column_dimensions["B"].width = 28
    ws_marks.column_dimensions["C"].width = 22
    ws_marks.column_dimensions["D"].width = 24
    ws_marks.column_dimensions["E"].width = 30

    # ----- Sheet 4: Grading Guide -----
    ws_guide = wb.create_sheet("Grading Guide")
    ws_guide.append(["UGANDA CBC GRADING SCALE (UNEB / NCDC)"])
    ws_guide.append([])
    ws_guide.append(["Score Range", "Grade", "Descriptor", "Meaning"])
    for low, high, letter, desc, full in GRADE_SCALE:
        ws_guide.append([f"{low} – {high}", letter, desc, full])
    ws_guide.append([])
    ws_guide.append(["Final Score Calculation"])
    ws_guide.append(["Final Mark = (Formative × 20%) + (Summative × 80%)"])
    ws_guide.append([])
    ws_guide.append(["Note: Project work / Activities of Integration are part of Formative assessment."])

    ws_guide["A1"].font = Font(bold=True, size=14)
    ws_guide.column_dimensions["A"].width = 15
    ws_guide.column_dimensions["B"].width = 10
    ws_guide.column_dimensions["C"].width = 15
    ws_guide.column_dimensions["D"].width = 90

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"✓ Excel template created: {output_path}")
    return output_path


# ============================================================
# PDF REPORT CARD GENERATOR
# ============================================================

def create_styles():
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name="SchoolName", fontName="Helvetica-Bold", fontSize=16,
        alignment=TA_CENTER, spaceAfter=2, textColor=colors.HexColor("#1F4E79")
    ))
    styles.add(ParagraphStyle(
        name="SchoolMotto", fontName="Helvetica-Oblique", fontSize=9,
        alignment=TA_CENTER, spaceAfter=4, textColor=colors.HexColor("#333333")
    ))
    styles.add(ParagraphStyle(
        name="SchoolContact", fontName="Helvetica", fontSize=8,
        alignment=TA_CENTER, spaceAfter=2, textColor=colors.HexColor("#555555")
    ))
    styles.add(ParagraphStyle(
        name="ReportTitle", fontName="Helvetica-Bold", fontSize=12,
        alignment=TA_CENTER, spaceBefore=6, spaceAfter=8,
        textColor=colors.HexColor("#1F4E79")
    ))
    styles.add(ParagraphStyle(
        name="SectionHeader", fontName="Helvetica-Bold", fontSize=10,
        alignment=TA_LEFT, spaceBefore=8, spaceAfter=4,
        textColor=colors.HexColor("#1F4E79")
    ))
    styles.add(ParagraphStyle(
        name="NormalSmall", fontName="Helvetica", fontSize=8,
        alignment=TA_LEFT, leading=11
    ))
    styles.add(ParagraphStyle(
        name="NormalTiny", fontName="Helvetica", fontSize=7,
        alignment=TA_LEFT, leading=9
    ))
    styles.add(ParagraphStyle(
        name="Comment", fontName="Helvetica", fontSize=8,
        alignment=TA_JUSTIFY, leading=11, spaceBefore=2
    ))
    styles.add(ParagraphStyle(
        name="Footer", fontName="Helvetica", fontSize=7,
        alignment=TA_CENTER, textColor=colors.HexColor("#666666")
    ))
    styles.add(ParagraphStyle(
        name="GradeA", fontName="Helvetica-Bold", fontSize=9,
        textColor=colors.HexColor("#006600")
    ))
    return styles


def build_report_card(
    student: Dict,
    marks: List[Dict],
    school: Dict,
    term: str,
    year: str,
    report_date: str,
    output_path: str
):
    """Generate a single PDF report card."""
    doc = SimpleDocTemplate(
        output_path,
        pagesize=A4,
        rightMargin=12 * mm,
        leftMargin=12 * mm,
        topMargin=10 * mm,
        bottomMargin=12 * mm
    )
    styles = create_styles()
    story = []

    # ----- Header -----
    story.append(Paragraph(school.get("name", "SECONDARY SCHOOL").upper(), styles["SchoolName"]))
    if school.get("motto"):
        story.append(Paragraph(f'"{school["motto"]}"', styles["SchoolMotto"]))
    contact_line = " | ".join(filter(None, [
        school.get("address"),
        school.get("phone"),
        school.get("email")
    ]))
    story.append(Paragraph(contact_line, styles["SchoolContact"]))
    story.append(Spacer(1, 3 * mm))
    story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor("#1F4E79")))
    story.append(Paragraph(
        f"END OF {term.upper()} REPORT CARD – ACADEMIC YEAR {year}",
        styles["ReportTitle"]
    ))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#1F4E79")))
    story.append(Spacer(1, 4 * mm))

    # ----- Learner Particulars -----
    story.append(Paragraph("LEARNER PARTICULARS", styles["SectionHeader"]))

    def clean(val):
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return ""
        return str(val).strip()

    full_name = " ".join(filter(None, [
        clean(student.get("Surname", "")),
        clean(student.get("First Name", "")),
        clean(student.get("Other Names", ""))
    ])).strip().upper()

    particulars = [
        [Paragraph("<b>Name:</b>", styles["NormalSmall"]),
         Paragraph(full_name, styles["NormalSmall"]),
         Paragraph("<b>Admission No:</b>", styles["NormalSmall"]),
         Paragraph(clean(student.get("Admission No", "")), styles["NormalSmall"])],
        [Paragraph("<b>Class:</b>", styles["NormalSmall"]),
         Paragraph(f"{clean(student.get('Class', ''))} {clean(student.get('Stream', ''))}", styles["NormalSmall"]),
         Paragraph("<b>Sex:</b>", styles["NormalSmall"]),
         Paragraph(clean(student.get("Sex", "")), styles["NormalSmall"])],
        [Paragraph("<b>Date of Birth:</b>", styles["NormalSmall"]),
         Paragraph(clean(student.get("Date of Birth", "")), styles["NormalSmall"]),
         Paragraph("<b>Parent/Guardian:</b>", styles["NormalSmall"]),
         Paragraph(clean(student.get("Parent/Guardian", "")), styles["NormalSmall"])],
        [Paragraph("<b>Contact:</b>", styles["NormalSmall"]),
         Paragraph(clean(student.get("Contact", "")), styles["NormalSmall"]),
         Paragraph("<b>Report Date:</b>", styles["NormalSmall"]),
         Paragraph(report_date, styles["NormalSmall"])],
    ]

    t = Table(particulars, colWidths=[28*mm, 55*mm, 32*mm, 55*mm])
    t.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F5F8FC")),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#1F4E79")),
    ]))
    story.append(t)
    story.append(Spacer(1, 5 * mm))

    # ----- Academic Performance -----
    story.append(Paragraph("ACADEMIC PERFORMANCE (Competency-Based Assessment)", styles["SectionHeader"]))
    story.append(Paragraph(
        "Final Score = (Formative × 20%) + (Summative × 80%) &nbsp;&nbsp;|&nbsp;&nbsp; "
        "Grading: A (Exceptional) · B (Outstanding) · C (Satisfactory) · D (Basic) · E (Elementary)",
        styles["NormalTiny"]
    ))
    story.append(Spacer(1, 2 * mm))

    # Table header
    header = [
        Paragraph("<b>Subject</b>", styles["NormalTiny"]),
        Paragraph("<b>Formative<br/>(20%)</b>", styles["NormalTiny"]),
        Paragraph("<b>Summative<br/>(80%)</b>", styles["NormalTiny"]),
        Paragraph("<b>Final<br/>Score</b>", styles["NormalTiny"]),
        Paragraph("<b>Grade</b>", styles["NormalTiny"]),
        Paragraph("<b>Descriptor</b>", styles["NormalTiny"]),
        Paragraph("<b>Teacher Comment</b>", styles["NormalTiny"]),
    ]

    data = [header]
    total_score = 0.0
    count = 0
    grade_counts = {"A": 0, "B": 0, "C": 0, "D": 0, "E": 0}

    for m in marks:
        form = m.get("Formative (out of 100)")
        summ = m.get("Summative (out of 100)")
        final = calculate_final_score(form, summ)
        letter, desc, _ = get_grade(final)

        if final is not None:
            total_score += final
            count += 1
            if letter in grade_counts:
                grade_counts[letter] += 1

        form_str = f"{float(form):.0f}" if not pd.isna(form) else "–"
        summ_str = f"{float(summ):.0f}" if not pd.isna(summ) else "–"
        final_str = f"{final:.1f}" if final is not None else "–"

        data.append([
            Paragraph(str(m.get("Subject", "")), styles["NormalTiny"]),
            Paragraph(form_str, styles["NormalTiny"]),
            Paragraph(summ_str, styles["NormalTiny"]),
            Paragraph(f"<b>{final_str}</b>", styles["NormalTiny"]),
            Paragraph(f"<b>{letter}</b>", styles["NormalTiny"]),
            Paragraph(desc, styles["NormalTiny"]),
            Paragraph(clean(m.get("Teacher Comment (optional)", "")), styles["NormalTiny"]),
        ])

    col_widths = [38*mm, 18*mm, 18*mm, 16*mm, 12*mm, 28*mm, 40*mm]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (1, 0), (4, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#AAAAAA")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F5FA")]),
        ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#1F4E79")),
    ]))
    story.append(table)
    story.append(Spacer(1, 4 * mm))

    # Summary box
    avg = round(total_score / count, 1) if count else 0
    avg_letter, avg_desc, _ = get_grade(avg)

    summary_data = [
        [Paragraph("<b>SUMMARY</b>", styles["NormalSmall"]),
         Paragraph(f"<b>Subjects Taken:</b> {count}", styles["NormalSmall"]),
         Paragraph(f"<b>Average Score:</b> {avg}", styles["NormalSmall"]),
         Paragraph(f"<b>Overall Grade:</b> {avg_letter} ({avg_desc})", styles["NormalSmall"])],
        [Paragraph(f"A: {grade_counts['A']} &nbsp; B: {grade_counts['B']} &nbsp; "
                   f"C: {grade_counts['C']} &nbsp; D: {grade_counts['D']} &nbsp; "
                   f"E: {grade_counts['E']}", styles["NormalTiny"]),
         "", "", ""],
    ]
    st = Table(summary_data, colWidths=[40*mm, 35*mm, 40*mm, 55*mm])
    st.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#E8F0FE")),
        ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#1F4E79")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("SPAN", (0, 1), (-1, 1)),
    ]))
    story.append(st)
    story.append(Spacer(1, 5 * mm))

    # ----- Attendance -----
    story.append(Paragraph("ATTENDANCE", styles["SectionHeader"]))
    att = [
        [Paragraph(f"<b>Days Present:</b> {clean(student.get('Days Present', '–'))}", styles["NormalSmall"]),
         Paragraph(f"<b>Days Absent:</b> {clean(student.get('Days Absent', '–'))}", styles["NormalSmall"]),
         Paragraph(f"<b>Attendance:</b> {clean(student.get('Attendance (%)', '–'))}%", styles["NormalSmall"])]
    ]
    at = Table(att, colWidths=[55*mm, 55*mm, 60*mm])
    at.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F5F8FC")),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#1F4E79")),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(at)
    story.append(Spacer(1, 4 * mm))

    # ----- Comments -----
    story.append(Paragraph("CLASS TEACHER'S COMMENT", styles["SectionHeader"]))
    ct_comment = clean(student.get("Class Teacher Comment", "")) or "No comment."
    story.append(Paragraph(ct_comment, styles["Comment"]))
    story.append(Spacer(1, 3 * mm))

    story.append(Paragraph("HEAD TEACHER'S COMMENT", styles["SectionHeader"]))
    ht_comment = clean(student.get("Head Teacher Comment", "")) or "No comment."
    story.append(Paragraph(ht_comment, styles["Comment"]))
    story.append(Spacer(1, 6 * mm))

    # ----- Signatures -----
    sig_data = [
        [Paragraph("_________________________", styles["NormalSmall"]),
         Paragraph("_________________________", styles["NormalSmall"]),
         Paragraph("_________________________", styles["NormalSmall"])],
        [Paragraph("<b>Class Teacher</b>", styles["NormalTiny"]),
         Paragraph("<b>Head Teacher</b>", styles["NormalTiny"]),
         Paragraph("<b>Parent / Guardian</b>", styles["NormalTiny"])],
        [Paragraph("Date: _______________", styles["NormalTiny"]),
         Paragraph("Date: _______________", styles["NormalTiny"]),
         Paragraph("Date: _______________", styles["NormalTiny"])],
    ]
    sig = Table(sig_data, colWidths=[60*mm, 60*mm, 50*mm])
    sig.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(sig)
    story.append(Spacer(1, 6 * mm))

    # ----- Grading Key (footer) -----
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#1F4E79")))
    story.append(Paragraph(
        "<b>GRADING KEY (UNEB/NCDC CBC):</b> "
        "A = Exceptional (80–100) · B = Outstanding (70–79) · C = Satisfactory (60–69) · "
        "D = Basic (50–59) · E = Elementary (0–49)",
        styles["Footer"]
    ))
    story.append(Paragraph(
        "This report is generated under the Competency-Based Curriculum. "
        "Project work and Activities of Integration form part of the Formative assessment.",
        styles["Footer"]
    ))
    story.append(Paragraph(
        f"Generated on {datetime.now().strftime('%d %B %Y at %H:%M')} | Uganda Secondary School Report Card System",
        styles["Footer"]
    ))

    doc.build(story)
    return output_path


# ============================================================
# MAIN PROCESSING
# ============================================================

def load_data(excel_path: str):
    """Load students, marks and settings from the Excel workbook."""
    xls = pd.ExcelFile(excel_path)

    # Settings
    settings_df = pd.read_excel(xls, sheet_name="School Settings", header=None)
    settings = {}
    for _, row in settings_df.iterrows():
        key = str(row[0]).strip() if pd.notna(row[0]) else ""
        val = str(row[1]).strip() if pd.notna(row[1]) else ""
        if key and val:
            settings[key] = val

    school = {
        "name": settings.get("School Name", SCHOOL_CONFIG["name"]),
        "motto": settings.get("Motto", SCHOOL_CONFIG["motto"]),
        "address": settings.get("Address", SCHOOL_CONFIG["address"]),
        "phone": settings.get("Phone", SCHOOL_CONFIG["phone"]),
        "email": settings.get("Email", SCHOOL_CONFIG["email"]),
        "district": settings.get("District", SCHOOL_CONFIG["district"]),
    }
    term = settings.get("Term", "Term 3")
    year = settings.get("Academic Year", "2025/2026")
    report_date = settings.get("Report Date", datetime.now().strftime("%d/%m/%Y"))

    students_df = pd.read_excel(xls, sheet_name="Students")
    marks_df = pd.read_excel(xls, sheet_name="Marks")

    return school, term, year, report_date, students_df, marks_df


def generate_all_reports(excel_path: str, output_dir: str = "generated_reports"):
    """Generate PDF report cards for every student in the Excel file."""
    Path(output_dir).mkdir(parents=True, exist_ok=True)

    school, term, year, report_date, students_df, marks_df = load_data(excel_path)

    print(f"\nSchool : {school['name']}")
    print(f"Term   : {term}  |  Year: {year}")
    print(f"Students found: {len(students_df)}")
    print("-" * 50)

    generated = []
    for _, student_row in students_df.iterrows():
        student = student_row.to_dict()
        adm = str(student.get("Admission No", "")).strip()
        if not adm or adm == "nan":
            continue

        student_marks = marks_df[marks_df["Admission No"].astype(str).str.strip() == adm]
        marks_list = student_marks.to_dict("records")

        surname = str(student.get("Surname", "Unknown")).strip()
        first = str(student.get("First Name", "")).strip()
        safe_adm = adm.replace("/", "-").replace("\\", "-")
        safe_name = f"{surname}_{first}".replace(" ", "_")
        filename = f"ReportCard_{safe_adm}_{safe_name}.pdf"
        out_path = os.path.join(output_dir, filename)

        build_report_card(
            student=student,
            marks=marks_list,
            school=school,
            term=term,
            year=year,
            report_date=report_date,
            output_path=out_path
        )
        print(f"  ✓ {filename}")
        generated.append(out_path)

    # Also create a simple class summary CSV
    summary_path = os.path.join(output_dir, "class_summary.csv")
    summary_rows = []
    for _, student_row in students_df.iterrows():
        student = student_row.to_dict()
        adm = str(student.get("Admission No", "")).strip()
        student_marks = marks_df[marks_df["Admission No"].astype(str).str.strip() == adm]
        scores = []
        for _, m in student_marks.iterrows():
            final = calculate_final_score(m.get("Formative (out of 100)"), m.get("Summative (out of 100)"))
            if final is not None:
                scores.append(final)
        avg = round(sum(scores) / len(scores), 1) if scores else None
        letter, desc, _ = get_grade(avg) if avg is not None else ("–", "", "")
        summary_rows.append({
            "Admission No": adm,
            "Name": f"{student.get('Surname', '')} {student.get('First Name', '')}",
            "Class": f"{student.get('Class', '')} {student.get('Stream', '')}",
            "Subjects": len(scores),
            "Average": avg,
            "Overall Grade": letter,
            "Descriptor": desc,
            "Attendance %": student.get("Attendance (%)"),
        })
    pd.DataFrame(summary_rows).to_csv(summary_path, index=False)
    print(f"\n✓ Class summary saved: {summary_path}")
    print(f"✓ Total report cards generated: {len(generated)}")
    return generated


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Uganda Secondary School Report Card System")
    parser.add_argument("--excel", default="templates/marks_entry_template.xlsx",
                        help="Path to the filled Excel marks file")
    parser.add_argument("--create-template", action="store_true",
                        help="Create a blank Excel template with sample data")
    parser.add_argument("--output", default="generated_reports",
                        help="Output directory for PDF report cards")
    args = parser.parse_args()

    if args.create_template:
        create_excel_template(args.excel)
        print("\nTemplate ready. Open it in Excel/LibreOffice, fill in real data, then run:")
        print(f"  python report_card_system.py --excel {args.excel}")
        return

    if not os.path.exists(args.excel):
        print(f"Excel file not found: {args.excel}")
        print("Creating a template with sample data first...")
        create_excel_template(args.excel)

    generate_all_reports(args.excel, args.output)


if __name__ == "__main__":
    main()
